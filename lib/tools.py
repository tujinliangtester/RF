import time
import xlrd, xlwt
import sys, os
from openpyxl import load_workbook


class tools(object):
    def __init__(self):
        self._sheet_name = 'Sheet3'
        self._max_line = 20

    def gene_mobile(self):
        s = time.strftime("%Y%m%d%H%M", time.localtime())
        s = s[-8:]
        s = '199' + s
        print(s)
        return s

    def read_data(self, key):
        path = os.path.join(os.path.dirname(__file__), '../config/setting.xlsx')
        tmp_dic = {}
        wb = xlrd.open_workbook(path)
        ws = wb.sheet_by_name(self._sheet_name)
        i = 2
        try:
            while (i < self._max_line):
                tmp_key = ws.cell_value(i, 1)
                if (tmp_key is None):
                    break
                tmp_dic[tmp_key] = ws.cell_value(i, 2)
                i += 1
        except Exception:
            print(Exception)

        return tmp_dic[key]

    # 这个函数还没有写完，这个函数只能新增文件，不能修改已存在文件
    def add_xlsx(self, s):
        path = os.path.join(os.path.dirname(__file__), '../config/setting.xlsx')
        tmp_dic = {}
        wb = xlwt.Workbook(path)
        ws = wb.sheet_index(self._sheet_name)

    def edit_xlsx(self, dic):
        path = os.path.join(os.path.dirname(__file__), '../config/setting.xlsx')
        wb = load_workbook(path)
        ws = wb[self._sheet_name]
        i = 3
        flag = False
        '''
        try:
            while (i < self._max_line):
                s = 'B' + str(i)
                for key in dic.keys():
                    if (ws[s].value is None):
                        break
                    if (ws[s].value == key):
                        s2 = 'C' + str(i)
                        ws[s2] = dic[key]
                        flag = True
                        break
                if (flag): break
                i += 1
        except Exception:
            print(Exception.mro())
            
            '''
        while (i < self._max_line):
            s = 'B' + str(i)
            for key in dic.keys():
                if (ws[s].value is None):
                    break
                if (ws[s].value == key):
                    s2 = 'C' + str(i)
                    ws[s2] = dic[key]
                    flag = True
                    break
            if (flag): break
            i += 1
        if (flag == False):  # 说明xlsx中不存在这个字段
            print(111)
            s = 'B' + str(i)
            ws[s] = dic.keys[0]
            s = 'C' + str(i)
            ws[s] = dic[dic.keys[0]]


if __name__ == '__main__':
    t = tools()
    # t.gene_mobile()
    print(t.read_data('mobile'))
    dic = {'mobile': 123}
    t.edit_xlsx(dic)
