import xlrd
import xlwt
import os
from openpyxl import load_workbook

_sheet_name='Sheet3'
_sheet_max_line = 20

def read_data( key):
    path = os.path.join(os.path.dirname(__file__), '../config/setting.xlsx')
    tmp_dic = {}
    wb = xlrd.open_workbook(path)
    ws = wb.sheet_by_name(_sheet_name)
    i = 2
    try:
        while (i < _sheet_max_line):
            tmp_key = ws.cell_value(i, 1)
            if (tmp_key is None):
                break
            tmp_dic[tmp_key] = ws.cell_value(i, 2)
            i += 1
    except Exception:
        erro='erro'
    print(tmp_dic[key])

    return tmp_dic[key]


# 这个函数还没有写完，这个函数只能新增文件，不能修改已存在文件
def add_xlsx( s):
    path = os.path.join(os.path.dirname(__file__), '../config/setting.xlsx')
    tmp_dic = {}
    wb = xlwt.Workbook(path)
    ws = wb.sheet_index(_sheet_name)


def edit_xlsx( dic):
    path = os.path.join(os.path.dirname(__file__), '../config/setting.xlsx')
    wb = load_workbook(path)
    ws = wb[_sheet_name]
    i = 3
    flag = False
    null_flag = False
    while (i < _sheet_max_line):
        s = 'B' + str(i)
        value = ws[s].value
        for key in dic.keys():
            if (ws[s].value is None):
                null_flag = True
                break
            if (ws[s].value == key):
                s2 = 'C' + str(i)
                ws[s2] = dic[key]
                flag = True
                break
        if (flag or null_flag): break
        i += 1
    if (flag == False):  # 说明xlsx中不存在这个字段
        s = 'B' + str(i)
        for key in dic.keys():
            ws[s] = key
            s = 'C' + str(i)
            ws[s] = dic[key]
    wb.save(path)
if __name__ == '__main__':

    mom='{"code":1,"msg":"","data":{"pointdistance":null,"mobile":null,"username":null,"nickname":null,"product_type_name":null,"site_name":null,"site_image":null,"gun_number":null,"longitude":0.0,"latitude":0.0,"is_disable":0,"id":10097194,"trade_id":97195,"merchant_id":61,"site_id":2,"user_id":27055,"order_no":"97195O10097194190318","trade_no":"T9719520190318155049837","is_online_pay":1,"pay_method_id":0,"is_oil":1,"fei_you_id":0,"gun_id":32,"gun_num":"10","oil_brand_id":"999       ","product_category":1,"product_type_id":12,"org_amt":10.00,"org_price":6.00,"price":6.00,"org_oil_litre":1.67,"site_activity_id":0,"site_activity_amt":0.00,"site_coupon_id":75,"site_coupon_amt":10.00,"pay_amt":0.00,"order_status":0,"create_time":"2019-03-18 15:50:49","pay_time":null,"cancel_time":null,"has_fei_you":0,"fei_you_amt":0.00,"coupon_nogas_amt":0.00,"coupon_nogas_id":0,"coin_amt":0.00,"score_amt":0.00,"site_activity_id2":0,"site_activity_amt2":0.00,"site_discount_type":0,"pay_status":0,"reduce_total_amt":0.00,"is_audit_sale":0,"audit_sale_user_id":null,"audit_sale_user":null,"audit_sale_date":null,"is_notify_pushed":0,"print_count":0,"pay_method_time":null,"pay_wx_orderno":null,"use_paymethod_type":0,"ticket_status":0,"pay_oil_amt":0.00,"pay_not_oil_amt":0.00,"pay_online_amt":0.00,"vehicle_pay_bill_id":0,"verify_status":0,"verify_time":null,"note":null},"ext":null,"subcode":""}'
    reg='"code":\d+'
    # t.reg_draw(mom=mom,key='id')
    res=read_data("mobile")
    print(res)
